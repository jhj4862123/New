import os
import shutil
import pyautogui
import openpyxl
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
from datetime import datetime
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tqdm import tqdm

############################ 이미지 폴더 선택 ########################################
root = Tk()
root.title("폴더 선택 창")   # 타이틀 설정

file_frame = Frame(root)
file_frame.pack(fill="x", padx = 5, pady= 5)

root.geometry("320x240") # 가로 *세로 사이즈
root.resizable(False, False)    #가로 *세로 사이즈 변경 가능 유무

dir_path = None        #폴더 경로 담을 변수 생성

def folder_select():
    global dir_path
    dir_path = filedialog.askdirectory(initialdir="./", \
                                       title = "폴더를 선택 해 주세요")  #folder 변수에 선택 폴더 경로 넣기
    if dir_path == '':
        messagebox.showwarning("경고", "폴더를 선택 하세요")    #폴더 선택 안했을 때 메세지 출력
    else:
        res = os.listdir(dir_path) # 폴더에 있는 파일 리스트 넣기
        if len(res) == 0:
            messagebox.showwarning("경고", "폴더내 파일이 없습니다.")
        else:
            root.destroy()

# def savefolder_select():
#     global save_dir_path
#     save_dir_path = filedialog.askdirectory(initialdir="./", \
#                                             title = "폴더를 선택 해 주세요")  #folder 변수에 선택 폴더 경로 넣기
#     if save_dir_path == '':
#         messagebox.showwarning("경고", "폴더를 선택 하세요")    #폴더 선택 안했을 때 메세지 출력
#     else:
#         res = os.listdir(save_dir_path) # 폴더에 있는 파일 리스트 넣기
#         if len(res) == 0:
#             messagebox.showwarning("경고", "폴더내 파일이 없습니다.")
#         else:
#             root.destroy()


btn_active_dir = Button(file_frame, text ="충전기 사진을 선택해 주세요. \n\n사진 형식 : 충전기번호_1.jpg\n ex) 1234_1.jpg", \
                        font=36, width = 24, padx = 10, pady= 20, command=folder_select)
btn_active_dir.pack( padx = 5, pady= 5)

# btn_active_dir = Button(file_frame, text ="파일을 저장할 폴더를 선택해 주세요. \n\n", \
#                         font=36, width = 24, padx = 10, pady= 20, command=savefolder_select())
# btn_active_dir.pack( padx = 5, pady= 5)

root.mainloop()

############################ 경로 및 양식 ########################################

now = datetime.now()
s = now.strftime("%Y-%m-%d")
finishpath = '완료폴더/'
newpath = finishpath + s

# photosrc = '작업 전 사진/'
photosrc = dir_path + '/'
movephoto = newpath + '/완료된 사진/'
move_resize_photo = newpath + '/축소 사진/'
movefilesrc = '완료폴더/'
path = '점검데이터.xlsx'
j = 1

print("\nphotosrc : ", photosrc)

if not os.path.exists(newpath):
    os.makedirs(newpath)

if not os.path.exists(movephoto):
    os.makedirs(movephoto)

if not os.path.exists(move_resize_photo):
    os.makedirs(move_resize_photo)

data = pd.read_excel(path)
base = photosrc
print("\nbase : ", base)

count_photo = [] # 사진의 갯수

############################# 충전기 갯수 카운트 및 이미지 리사이즈 ########################################
from PIL import Image

chargernum = 0 # 충전기의 갯수

for name in data.iloc[0,1:]: #None 없애기
    chargernum += 1
    k = 0

    for j in range(1, 7):
        fileName = os.path.join(base, str(name) + "_" + str(j) + ".jpg")
        tempName = os.path.join(base, str(name) + "-" + str(j) + ".jpg")


        if os.path.exists(fileName):
            img = Image.open(fileName)
        elif os.path.exists(tempName):
            shutil.move(tempName, fileName)
            img = Image.open(fileName)
        else:
            #print(f"{fileName}이 없습니다.")
            k += 1
            globals()['test' + str(chargernum)] = k
            # print('test{}:'.format(chargernum))
            NoImgCount = 'test{}'.format(chargernum)
            # print(globals()[f'{NoImgCount}'])
            continue
        img = img.convert('RGB')
        resize_img = img.resize((584, 378))
        resize_img.save(base + str(name) + "_" + str(j) + "(resize).jpg")

############################# 양식 ########################################

from openpyxl.drawing.image import Image

wbMaster = load_workbook('점검양식.xlsx')
wsMaster = wbMaster['점검양식']
wbSlave = load_workbook('점검데이터.xlsx', data_only=True)
slavestandard = wbSlave['참조데이터']
wsSlave = wbSlave['점검정보']

############################# 변수들 ########################################

for i in tqdm(range(chargernum)):

    #print("===> 사진이 없는 갯수 : ")
    NoImgCount = 'test{}'.format(i+1)
    #print(globals()[f'{NoImgCount}'])

    if globals()[f'{NoImgCount}'] == 6:
        continue

    wbMaster = load_workbook('점검양식.xlsx')
    wsMaster = wbMaster['점검양식']
    # wsMaster = wbMaster.active
    slavestandard = wbSlave['참조데이터']
    wsSlave = wbSlave['점검정보']

    copynum = wsSlave['2'][i + 1].value  # 충전기 번호
    wsMaster['G7'] = copynum
    wsMaster['G7'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    print(f"충전소 이름이 출력되었습니다.{slavestandard[2][1+i].value}")
    wsMaster['C7'] = slavestandard[2][1+i].value
    wsMaster['C7'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    print(f"충전기 제조사가 출력되었습니다.{slavestandard[3][1+i].value}")
    wsMaster['C8'] = slavestandard[3][1+i].value
    wsMaster['C8'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    print(f"충전소 좌표가 출력되었습니다.{slavestandard[4][1+i].value}")
    wsMaster['C9'] = slavestandard[4][1+i].value
    wsMaster['C9'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    print(f"충전소 주소가 출력되었습니다.{slavestandard[5][1+i].value}")
    wsMaster['C10'] = slavestandard[5][1+i].value
    wsMaster['C10'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    print(f"위치가 출력되었습니다.{wsSlave[29][1].value}")

    if wsSlave[29][1 + i].value is None: # 변경하지 않은 값
        wsMaster['D13'] = slavestandard[6][1+i].value
        wsMaster['D13'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    else: # 변경한 값
        wsMaster['D13'] = wsSlave[29][1 + i].value
        wsMaster['D13'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    #    print(f"충전기 용량이 출력되었습니다.{slavestandard[7][1+i].value}")
    wsMaster['G15'] = slavestandard[7][1+i].value
    wsMaster['G15'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyname = wsSlave['3'][i+1].value  # 점검자 이름
    wsMaster['G3'] = copyname
    wsMaster['G3'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyday = wsSlave['4'][i+1].value  # 점검 날짜
    wsMaster['C3'] = copyday
    wsMaster['C3'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    day001 = pd.to_datetime(copyday, format='%Y-%m-%d')
    day001 = day001.date()
    #    print(day001)

    copycount = wsSlave['5'][i + 1].value  # 수량
    wsMaster['G9'] = copycount
    wsMaster['G9'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copytemp = wsSlave['6'][i + 1].value  # 온도
    wsMaster['C4'] = copytemp
    wsMaster['G25'] = copytemp
    wsMaster['C4'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    wsMaster['G25'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyhumi = wsSlave['7'][i + 1].value  # 습도
    wsMaster['G4'] = copyhumi
    wsMaster['G4'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    if wsSlave[8][1 + i].value is None: # 충전기 설치유형
        wsMaster['C11'] = ("벽걸이형")
        wsMaster['C11'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    else: # 변경한 값
        wsMaster['C11'] = wsSlave[8][1 + i].value
        wsMaster['C11'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyvolt = wsSlave['11'][i + 1].value  # 전압
    wsMaster['G14'] = copyvolt
    wsMaster['G14'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copylux = wsSlave['12'][i + 1].value  # 조도
    wsMaster['G36'] = copylux
    wsMaster['G36'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyer = wsSlave['13'][i + 1].value  # 접지저항
    wsMaster['G65'] = copyer
    wsMaster['G65'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyir = wsSlave['14'][i + 1].value  # 절연저항
    wsMaster['G69'] = copyir
    wsMaster['G69'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copymb = wsSlave['15'][i + 1].value  # 메인차단
    wsMaster['G59'] = copymb
    wsMaster['G59'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copylb = wsSlave['16'][i + 1].value  # 누전차단
    wsMaster['G60'] = copylb
    wsMaster['G60'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copysc = wsSlave['17'][i + 1].value  # 감도전류
    wsMaster['G61'] = copysc
    wsMaster['G61'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyes = wsSlave['18'][i + 1].value  # 비상정지
    wsMaster['G42'] = copyes
    wsMaster['G42'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copystopper = wsSlave['20'][i + 1].value  # 스토퍼
    wsMaster['G38'] = copystopper
    wsMaster['G38'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyil = wsSlave['23'][i + 1].value  # 설치위치
    # wsMaster[#위치불명] = copyil
    wsMaster['G42'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyesyn = wsSlave['24'][i + 1].value  # 비상정지
    wsMaster['G79'] = copyesyn
    wsMaster['G79'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copyfa = wsSlave['25'][i + 1].value  # 소화기 비치
    wsMaster['G71'] = copyfa
    wsMaster['G71'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    copysp = wsSlave['26'][i + 1].value  # 스프링쿨러
    wsMaster['G70'] = copysp
    wsMaster['G70'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')

    if wsSlave[27][1 + i].value != None: # 충전기 설치유형
        wsMaster['C81'] = wsSlave[27][1 + i].value
        wsMaster['C81'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    ############################ 사진 ########################################

    # 사진이 하나도 없으면 이 For문을 돌지 않게

    for j in range(1, 7):
        src_img_1 = os.path.join(base, str(copynum) + "_1.jpg")
        src_img_2 = os.path.join(base, str(copynum) + "_2.jpg")
        src_img_3 = os.path.join(base, str(copynum) + "_3.jpg")
        src_img_4 = os.path.join(base, str(copynum) + "_4.jpg")
        src_img_5 = os.path.join(base, str(copynum) + "_5.jpg")
        src_img_6 = os.path.join(base, str(copynum) + "_6.jpg")

        if os.path.exists(src_img_1):
            fileName = os.path.join(base, str(copynum) + "_1(resize).jpg")
            if os.path.exists(fileName):
                img1 = Image(fileName)
                wsMaster.add_image(img1, 'A84')
                shutil.move(photosrc + str(copynum) + "_1.jpg", movephoto + str(copynum) + "_1.jpg")
        if os.path.exists(src_img_2):
            fileName = os.path.join(base, str(copynum) + "_2(resize).jpg")
            if os.path.exists(fileName):
                img2 = Image(fileName)
                wsMaster.add_image(img2, 'F84')
                shutil.move(photosrc + str(copynum) + "_2.jpg", movephoto + str(copynum) + "_2.jpg")
        if os.path.exists(src_img_3):
            fileName = os.path.join(base, str(copynum) + "_3(resize).jpg")
            if os.path.exists(fileName):
                img3 = Image(fileName)
                wsMaster.add_image(img3, 'A103')
                shutil.move(photosrc + str(copynum) + "_3.jpg", movephoto + str(copynum) + "_3.jpg")
        if os.path.exists(src_img_4):
            fileName = os.path.join(base, str(copynum) + "_4(resize).jpg")
            if os.path.exists(fileName):
                img4 = Image(fileName)
                wsMaster.add_image(img4, 'F103')
                shutil.move(photosrc + str(copynum) + "_4.jpg", movephoto + str(copynum) + "_4.jpg")
        if os.path.exists(src_img_5):
            fileName = os.path.join(base, str(copynum) + "_5(resize).jpg")
            if os.path.exists(fileName):
                img5 = Image(fileName)
                wsMaster.add_image(img5, 'A122')
                shutil.move(photosrc + str(copynum) + "_5.jpg", movephoto + str(copynum) + "_5.jpg")
        if os.path.exists(src_img_6):
            fileName = os.path.join(base, str(copynum) + "_6(resize).jpg")
            if os.path.exists(fileName):
                img6 = Image(fileName)
                wsMaster.add_image(img6, 'F122')
                shutil.move(photosrc + str(copynum) + "_6.jpg", movephoto + str(copynum) + "_6.jpg")

    ############################# 출력형식 ########################################

    if globals()[f'{NoImgCount}'] != 6:
        wbMaster.save(str(copynum) + "_" + str(copyname) + "_" + str(day001) + ".xlsx")
    else:
        continue
    shutil.move(str(copynum) + "_" + str(copyname) + "_" + str(day001) + ".xlsx",
                newpath + "/" + str(copynum) + "_" + str(copyname) + "_" + str(day001) + ".xlsx")
    wbMaster.close()
    print(str(copynum) + "_" + str(copyname) + "_" + str(day001) + ".xlsx" + " 파일이 생성되었습니다.")


file_list = os.listdir(base) # 폴더안의 파일 리스트를 얻습니다.

for item in file_list:
    if item[-12:] == "(resize).jpg": # item[-12:] 마지막 12글자
        shutil.move(photosrc + item, move_resize_photo + item)

print("총" + str(chargernum) + "개의 파일이 생성되었습니다.")
input("엔터를 누르면 종료됩니다.")
exit()